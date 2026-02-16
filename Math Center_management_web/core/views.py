from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User, Group
from django.db.models import Count, Sum, Avg, Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .models import nhan_vien, teacher, hoc_vien, clazz, schedule, enrollments, attendance, feedback, class_type
from django.contrib import messages
from django.utils import timezone
from datetime import datetime, timedelta
from django.shortcuts import render, redirect, get_object_or_404
from .models import teacher
from .forms import TeacherForm, HocVienForm, ClassForm, NhanVienForm, ClassTypeForm, ScheduleForm


def is_admin(user):
    return user.is_authenticated and (user.is_superuser or user.is_staff)

# Home View
def home(request):
    # Luôn chuyển hướng về login
    return redirect('core:login')

# Authentication Views
def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            # Kiểm tra user có phải admin không
            if user.is_superuser or user.is_staff:
                login(request, user)
                return redirect('core:admin_dashboard')
            else:
                messages.error(request, 'Bạn không có quyền truy cập hệ thống này')
        else:
            messages.error(request, 'Tên đăng nhập hoặc mật khẩu không đúng')
    return render(request, 'core/login.html')

@login_required
def logout_view(request):
    logout(request)
    return redirect('core:login')

# Dashboard Views
@login_required
@user_passes_test(is_admin, login_url='core:login')
def admin_dashboard(request):
    from django.db import connection
    from datetime import date
    
    with connection.cursor() as cursor:
        # 1. Doanh thu theo từng tháng của năm hiện tại
        current_year = date.today().year
        cursor.execute("""
            SELECT 
                EXTRACT(MONTH FROM e.enrollment_date) as month,
                SUM(c.price) as revenue
            FROM enrollments e
            JOIN clazz c ON e.class_id = c.class_id
            WHERE EXTRACT(YEAR FROM e.enrollment_date) = %s
            GROUP BY EXTRACT(MONTH FROM e.enrollment_date)
            ORDER BY month;
        """, [current_year])
        monthly_revenue_current = cursor.fetchall()
        

        
        # 2. Số lượng lớp đang mở hiện tại
        cursor.execute("""
            SELECT COUNT(*) 
            FROM clazz 
            WHERE khai_giang <= %s AND ket_thuc >= %s;
        """, [date.today(), date.today()])
        active_classes = cursor.fetchone()[0]
        
        # 3. Số lượng nhân viên
        cursor.execute("SELECT COUNT(*) FROM nhan_vien;")
        total_staff = cursor.fetchone()[0]
        
        # 4. Số lượng giáo viên
        cursor.execute("SELECT COUNT(*) FROM teacher;")
        total_teachers = cursor.fetchone()[0]
        
        # 5. Số lượng học viên đang học (đang học lớp còn mở)
        cursor.execute("""
            SELECT COUNT(DISTINCT e.student_id)
            FROM enrollments e
            JOIN clazz c ON e.class_id = c.class_id
            WHERE c.khai_giang <= %s AND c.ket_thuc >= %s;
        """, [date.today(), date.today()])
        active_students = cursor.fetchone()[0]
        
        # 6. Bảng xếp hạng thống kê đánh giá theo từng loại lớp
        cursor.execute("""
            SELECT 
                ct.describe as class_type_name,
                ct.code as class_type_code,
                COUNT(f.id_feedback) as total_feedbacks,
                ROUND(AVG(f.class_rate), 2) as avg_class_rating,
                ROUND(AVG(f.teacher_rate), 2) as avg_teacher_rating
            FROM class_type ct
            LEFT JOIN clazz c ON ct.type_id = c.type_id
            LEFT JOIN feedback f ON c.class_id = f.class_id
            GROUP BY ct.type_id, ct.describe, ct.code
            ORDER BY avg_class_rating DESC, avg_teacher_rating DESC;
        """)
        class_type_ratings = cursor.fetchall()
        
        # 7. Bảng xếp hạng top 10 điểm đánh giá giáo viên
        cursor.execute("""
            SELECT 
                t.full_name as teacher_name,
                t.teacher_id,
                COUNT(f.id_feedback) as total_feedbacks,
                ROUND(AVG(f.teacher_rate), 2) as avg_teacher_rating
            FROM teacher t
            LEFT JOIN feedback f ON t.teacher_id = f.teacher_id
            GROUP BY t.teacher_id, t.full_name
            HAVING COUNT(f.id_feedback) > 0
            ORDER BY avg_teacher_rating DESC, total_feedbacks DESC
            LIMIT 10;
        """)
        teacher_rankings = cursor.fetchall()
        
        # 8. Bảng xếp hạng top 10 điểm số tổng kết học viên
        cursor.execute("""
    SELECT * FROM top_10_student_final_scores
    """)
        student_rankings = cursor.fetchall()
    
    # Chuẩn bị dữ liệu doanh thu theo tháng (12 tháng)
    revenue_data = [0] * 12  # Khởi tạo 12 tháng với doanh thu = 0
    for month, revenue in monthly_revenue_current:
        revenue_data[int(month) - 1] = int(revenue) if revenue else 0
    
    import json
    
    # Đảm bảo dữ liệu safe
    months_list = ['Tháng 1', 'Tháng 2', 'Tháng 3', 'Tháng 4', 'Tháng 5', 'Tháng 6',
                   'Tháng 7', 'Tháng 8', 'Tháng 9', 'Tháng 10', 'Tháng 11', 'Tháng 12']
    
    
    context = {
        'total_staff': total_staff,
        'total_teachers': total_teachers,
        'active_classes': active_classes,
        'active_students': active_students,
        'current_year': current_year,
        'monthly_revenue_current_json': json.dumps(revenue_data, ensure_ascii=False),
        'months_json': json.dumps(months_list, ensure_ascii=False),
        'class_type_ratings': class_type_ratings,
        'teacher_rankings': teacher_rankings,
        'student_rankings': student_rankings,
    }
    return render(request, 'core/admin_dashboard.html', context)

@login_required
def statistics(request):
    context = {
        'student_stats': hoc_vien.objects.aggregate(
            total=Count('student_id'),
        ),
        'teacher_stats': teacher.objects.aggregate(
            total=Count('teacher_id'),
        ),
        'class_stats':clazz.objects.aggregate(
            total=Count('class_id'),
        ),
        'attendance_stats': attendance.objects.aggregate(
            total=Count('id_attend'),
            present=Count('id_attend', filter=Q(status='Có mặt'))
        ),
        'feedback_stats': feedback.objects.aggregate(
            total=Count('id_feedback'),
            average_class_rating=Avg('class_rate'),
            average_teacher_rating=Avg('teacher_rate')
        ),
    }
    return render(request, 'core/statistics.html', context)



# Student Management Views
@login_required
def student_list(request):
    search_query = request.GET.get('search', '')
    students = hoc_vien.objects.all()
    
    if search_query:
        students = students.filter(
            Q(full_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(sdt__icontains=search_query) |
            Q(student_id__icontains=search_query)
        )
    
    # Phân trang - 10 học viên mỗi trang
    paginator = Paginator(students, 10)
    page = request.GET.get('page')
    
    try:
        students_page = paginator.page(page)
    except PageNotAnInteger:
        students_page = paginator.page(1)
    except EmptyPage:
        students_page = paginator.page(paginator.num_pages)
    
    context = {
        'students': students_page,
        'search_query': search_query,
    }
    return render(request, 'students/student_list.html', context)

@login_required
def student_create(request):
    if request.method == 'POST':
        form = HocVienForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Thêm học viên thành công!')
            return redirect('core:student_list')
    else:
        form = HocVienForm()
    return render(request, 'students/student_form.html', {'form': form, 'action': 'Thêm'})

@login_required
def student_edit(request, pk):
    student = get_object_or_404(hoc_vien, pk=pk)
    if request.method == 'POST':
        form = HocVienForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cập nhật học viên thành công!')
            return redirect('core:student_list')
    else:
        form = HocVienForm(instance=student)
    return render(request, 'students/student_form.html', {'form': form, 'action': 'Sửa'})

@login_required
def student_delete(request, pk):
    student = get_object_or_404(hoc_vien, pk=pk)
    if request.method == 'POST':
        student.delete()
        messages.success(request, 'Xóa học viên thành công!')
        return redirect('core:student_list')
    return render(request, 'students/student_confirm_delete.html', {'student': student})

#-------------------------------
# Teacher Management Views
#-------------------------------
@login_required
def teacher_list(request):
    search_query = request.GET.get('search', '')
    teachers = teacher.objects.all()
    
    if search_query:
        teachers = teachers.filter(
            Q(full_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(sdt__icontains=search_query) |
            Q(teacher_id__icontains=search_query) |
            Q(trinh_do__icontains=search_query)
        )
    
    # Phân trang - 10 giáo viên mỗi trang
    paginator = Paginator(teachers, 10)
    page = request.GET.get('page')
    
    try:
        teachers_page = paginator.page(page)
    except PageNotAnInteger:
        teachers_page = paginator.page(1)
    except EmptyPage:
        teachers_page = paginator.page(paginator.num_pages)
    
    context = {
        'teachers': teachers_page,
        'search_query': search_query,
    }
    return render(request, 'teachers/teacher_list.html', context)

@login_required
def teacher_create(request):
    if request.method == 'POST':
        form = TeacherForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('core:teacher_list')
    else:
        form = TeacherForm()
    return render(request, 'core/teacher_form.html', {'form': form, 'action': 'Thêm'})
@login_required
def teacher_edit(request, pk):
    gv = get_object_or_404(teacher, pk=pk)
    if request.method == 'POST':
        form = TeacherForm(request.POST, instance=gv)
        if form.is_valid():
            form.save()
            return redirect('core:teacher_list')
    else:
        form = TeacherForm(instance=gv)
    return render(request, 'core/teacher_form.html', {'form': form, 'action': 'Sửa'})
@login_required
def teacher_delete(request, pk):
    gv = get_object_or_404(teacher, pk=pk)
    if request.method == 'POST':
        gv.delete()
        return redirect('core:teacher_list')
    return render(request, 'core/teacher_confirm_delete.html', {'teacher': gv})

#-------------------------------
# Class Management Views
#-------------------------------
@login_required
def class_list(request):
    search_query = request.GET.get('search', '')
    classes = clazz.objects.select_related('teacher', 'nhan_vien', 'type').all()
    
    if search_query:
        classes = classes.filter(
            Q(class_name__icontains=search_query) |
            Q(teacher__full_name__icontains=search_query) |
            Q(nhan_vien__full_name__icontains=search_query) |
            Q(type__describe__icontains=search_query) |
            Q(class_id__icontains=search_query) |
            Q(room__icontains=search_query)
        )
    
    # Phân trang - 10 lớp học mỗi trang
    paginator = Paginator(classes, 10)
    page = request.GET.get('page')
    
    try:
        classes_page = paginator.page(page)
    except PageNotAnInteger:
        classes_page = paginator.page(1)
    except EmptyPage:
        classes_page = paginator.page(paginator.num_pages)
    
    context = {
        'classes': classes_page,
        'search_query': search_query,
    }
    return render(request, 'classes/class_list.html', context)

@login_required
def class_create(request):
    print("POST DATA:", request.POST)
    if request.method == 'POST':
           form = ClassForm(request.POST)
           if form.is_valid():
               form.save()
               messages.success(request, 'Thêm lớp học thành công!')
               return redirect('core:class_list')
           # Nếu form không hợp lệ, sẽ render lại form với lỗi
    else:
           form = ClassForm()
    return render(request, 'core/class_form.html', {'form': form, 'action': 'Thêm'})

@login_required
def class_edit(request, pk):
    class_obj = get_object_or_404(clazz, pk=pk)
    if request.method == 'POST':
        form = ClassForm(request.POST, instance=class_obj)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cập nhật lớp học thành công!')
            return redirect('core:class_list')
    else:
        form = ClassForm(instance=class_obj)
    return render(request, 'core/class_form.html', {'form': form, 'action': 'Sửa'})

@login_required
def class_delete(request, pk):
    class_obj = get_object_or_404(clazz, pk=pk)
    if request.method == 'POST':
        class_obj.delete()
        return redirect('core:class_list')
    return render(request, 'classes/class_confirm_delete.html', {'class_obj': class_obj})

@login_required
def class_detail(request, pk):
    class_obj = get_object_or_404(clazz, pk=pk)
    enrolls = enrollments.objects.filter(class_obj=class_obj)
    students = hoc_vien.objects.exclude(enrollments__class_obj=class_obj)

    # Lấy lịch học hiện tại (nếu có)
    try:
        current_schedule = class_obj.schedule
    except:
        current_schedule = None

    if request.method == 'POST':
        # Xử lý thêm học viên
        if 'student_id' in request.POST:
            student_id = request.POST.get('student_id')
            if student_id:
                student = get_object_or_404(hoc_vien, pk=student_id)
                # Kiểm tra sĩ số
                if enrolls.count() < 30:  
                    enrollments.objects.create(
                        student=student,
                        class_obj=class_obj,
                        enrollment_date=timezone.now()
                    )
                    messages.success(request, 'Thêm học viên vào lớp thành công!')
                else:
                    messages.error(request, 'Lớp đã đủ sĩ số!')
        
        # Xử lý xóa học viên
        elif 'remove_student' in request.POST:
            student_id = request.POST.get('remove_student')
            if student_id:
                enrollment = get_object_or_404(enrollments, student_id=student_id, class_obj=class_obj)
                enrollment.delete()
                messages.success(request, 'Xóa học viên khỏi lớp thành công!')

        # Xử lý cập nhật điểm số
        elif 'update_scores' in request.POST:
            student_id = request.POST.get('update_scores')
            if student_id:
                enrollment = get_object_or_404(enrollments, student_id=student_id, class_obj=class_obj)
                # Cập nhật điểm số
                enrollment.minitest1 = request.POST.get('minitest1') or None
                enrollment.minitest2 = request.POST.get('minitest2') or None
                enrollment.minitest3 = request.POST.get('minitest3') or None
                enrollment.minitest4 = request.POST.get('minitest4') or None
                enrollment.midterm = request.POST.get('midterm') or None
                enrollment.final_test = request.POST.get('final_test') or None
                enrollment.save()
                messages.success(request, 'Cập nhật điểm số thành công!')

        # Xử lý thêm/cập nhật lịch học
        elif 'add_schedule' in request.POST:
            if current_schedule:
                schedule_form = ScheduleForm(request.POST, instance=current_schedule)
            else:
                schedule_form = ScheduleForm(request.POST)
            
            if schedule_form.is_valid():
                schedule_obj = schedule_form.save(commit=False)
                schedule_obj.class_obj = class_obj
                schedule_obj.save()
                messages.success(request, 'Cập nhật lịch học thành công!')
                return redirect('core:class_detail', pk=pk)
            else:
                messages.error(request, 'Có lỗi khi thêm lịch học. Vui lòng kiểm tra lại!')
        
        return redirect('core:class_detail', pk=pk)
    
    # Tạo form lịch học
    if current_schedule:
        schedule_form = ScheduleForm(instance=current_schedule)
    else:
        schedule_form = ScheduleForm()

    return render(request, 'classes/class_detail.html', {
        'class_obj': class_obj,
        'enrolls': enrolls,
        'students': students,
        'current_schedule': current_schedule,
        'schedule_form': schedule_form,
    })

#-------------------------------
# Schedule Management Views
#-------------------------------
@login_required
def schedule_list(request):
    schedules = schedule.objects.all()
    
    # Phân trang - 10 lịch học mỗi trang
    paginator = Paginator(schedules, 10)
    page = request.GET.get('page')
    
    try:
        schedules_page = paginator.page(page)
    except PageNotAnInteger:
        schedules_page = paginator.page(1)
    except EmptyPage:
        schedules_page = paginator.page(paginator.num_pages)
    
    return render(request, 'core/schedule_list.html', {'schedules': schedules_page})

@login_required
def schedule_create(request):
    if request.method == 'POST':
        # Handle schedule creation
        pass
    return render(request, 'core/schedule_form.html')

@login_required
def schedule_edit(request, pk):
    schedule = get_object_or_404(schedule, pk=pk)
    if request.method == 'POST':
        # Handle schedule update
        pass
    return render(request, 'core/schedule_form.html', {'schedule': schedule})

@login_required
def schedule_delete(request, pk):
    schedule = get_object_or_404(schedule, pk=pk)
    if request.method == 'POST':
        schedule.delete()
        messages.success(request, 'Lịch học đã được xóa thành công')
        return redirect('core:schedule_list')
    return render(request, 'core/schedule_confirm_delete.html', {'schedule': schedule})

# Attendance Management Views
@login_required
def attendance_list(request):
    attendance_records = attendance.objects.all()
    
    # Phân trang - 15 bản ghi điểm danh mỗi trang
    paginator = Paginator(attendance_records, 15)
    page = request.GET.get('page')
    
    try:
        attendance_page = paginator.page(page)
    except PageNotAnInteger:
        attendance_page = paginator.page(1)
    except EmptyPage:
        attendance_page = paginator.page(paginator.num_pages)
    
    return render(request, 'core/attendance_list.html', {'attendance_records': attendance_page})

@login_required
def attendance_create(request):
    if request.method == 'POST':
        # Handle attendance creation
        pass
    return render(request, 'core/attendance_form.html')

@login_required
def attendance_edit(request, pk):
    attendance = get_object_or_404(attendance, pk=pk)
    if request.method == 'POST':
        # Handle attendance update
        pass
    return render(request, 'core/attendance_form.html', {'attendance': attendance})

@login_required
def attendance_delete(request, pk):
    attendance = get_object_or_404(attendance, pk=pk)
    if request.method == 'POST':
        attendance.delete()
        messages.success(request, 'Điểm danh đã được xóa thành công')
        return redirect('core:attendance_list')
    return render(request, 'core/attendance_confirm_delete.html', {'attendance': attendance})

# Feedback Management Views
@login_required
def feedback_list(request):
    feedbacks = feedback.objects.all().order_by('-id_feedback')
    students = hoc_vien.objects.all()
    teachers = teacher.objects.all()
    classes = clazz.objects.all()
    
    # Phân trang - 15 đánh giá mỗi trang
    paginator = Paginator(feedbacks, 15)
    page = request.GET.get('page')
    
    try:
        feedbacks_page = paginator.page(page)
    except PageNotAnInteger:
        feedbacks_page = paginator.page(1)
    except EmptyPage:
        feedbacks_page = paginator.page(paginator.num_pages)
    
    return render(request, 'core/feedback_list.html', {
        'feedbacks': feedbacks_page,
        'students': students,
        'teachers': teachers,
        'classes': classes,
    })

@login_required
def feedback_create(request):
    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        teacher_id = request.POST.get('teacher_id')
        class_id = request.POST.get('class_id')
        teacher_rate = request.POST.get('teacher_rate')
        class_rate = request.POST.get('class_rate')
        
        try:
            student = get_object_or_404(hoc_vien, student_id=student_id)
            teacher_obj = get_object_or_404(teacher, teacher_id=teacher_id)
            class_obj = get_object_or_404(clazz, class_id=class_id)
            
            feedback.objects.create(
                student=student,
                teacher=teacher_obj,
                class_obj=class_obj,
                teacher_rate=teacher_rate,
                class_rate=class_rate
            )
            messages.success(request, 'Thêm đánh giá thành công!')
        except Exception as e:
            messages.error(request, f'Có lỗi xảy ra: {str(e)}')
    
    return redirect('core:feedback_list')

@login_required
def feedback_import(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        from decimal import Decimal
        import traceback
        import openpyxl
        
        try:
            excel_file = request.FILES['excel_file']
            skip_header = request.POST.get('skip_header')
            
            print(f"File name: {excel_file.name}")
            print(f"File size: {excel_file.size}")
            print(f"Skip header: {skip_header}")
            
            # Chỉ xử lý file Excel (.xlsx)
            if not excel_file.name.endswith('.xlsx'):
                messages.error(request, 'Chỉ hỗ trợ file Excel (.xlsx). Vui lòng chuyển đổi file của bạn sang định dạng .xlsx')
                return redirect('core:feedback_list')
            
            # Đọc file Excel bằng openpyxl
            try:
                workbook = openpyxl.load_workbook(excel_file)
                worksheet = workbook.active
                print(f"Successfully loaded Excel file with {worksheet.max_row} rows and {worksheet.max_column} columns")
            except Exception as e:
                messages.error(request, f'Không thể đọc file Excel: {str(e)}')
                return redirect('core:feedback_list')
            
            success_count = 0
            error_count = 0
            errors = []
            
            # Xác định dòng bắt đầu
            start_row = 2 if skip_header else 1
            
            print(f"Processing from row {start_row} to {worksheet.max_row}")
            
            for row_num in range(start_row, worksheet.max_row + 1):
                try:
                    # Đọc dữ liệu từ các cột
                    # Cột A: Timestamp (bỏ qua)
                    # Cột B: Mã học viên
                    # Cột C: Mã giáo viên  
                    # Cột D: Mã lớp học
                    # Cột E: Điểm giáo viên
                    # Cột F: Điểm lớp học
                    
                    timestamp = worksheet.cell(row=row_num, column=1).value
                    student_id_raw = worksheet.cell(row=row_num, column=2).value
                    teacher_id_raw = worksheet.cell(row=row_num, column=3).value
                    class_id_raw = worksheet.cell(row=row_num, column=4).value
                    teacher_rate_raw = worksheet.cell(row=row_num, column=5).value
                    class_rate_raw = worksheet.cell(row=row_num, column=6).value
                    
                    print(f"Row {row_num}: timestamp={timestamp}, student_id={student_id_raw}, teacher_id={teacher_id_raw}, class_id={class_id_raw}, teacher_rate={teacher_rate_raw}, class_rate={class_rate_raw}")
                    
                    # Kiểm tra dòng trống
                    if not any([student_id_raw, teacher_id_raw, class_id_raw, teacher_rate_raw, class_rate_raw]):
                        print(f"Skipping empty row {row_num}")
                        continue
                    
                    # Kiểm tra dữ liệu bị thiếu
                    if not all([student_id_raw, teacher_id_raw, class_id_raw, teacher_rate_raw, class_rate_raw]):
                        errors.append(f"Dòng {row_num}: Thiếu dữ liệu bắt buộc")
                        error_count += 1
                        continue
                    
                    # Chuyển đổi kiểu dữ liệu
                    try:
                        student_id = int(float(student_id_raw))
                        teacher_id = int(float(teacher_id_raw))
                        class_id = int(float(class_id_raw))
                        teacher_rate = int(float(teacher_rate_raw))
                        class_rate = int(float(class_rate_raw))
                    except (ValueError, TypeError) as e:
                        errors.append(f"Dòng {row_num}: Lỗi chuyển đổi dữ liệu - {str(e)}")
                        error_count += 1
                        continue
                    
                    print(f"Parsed data: student_id={student_id}, teacher_id={teacher_id}, class_id={class_id}, teacher_rate={teacher_rate}, class_rate={class_rate}")
                    
                    # Validate điểm số
                    if not (1 <= teacher_rate <= 10):
                        errors.append(f"Dòng {row_num}: Điểm giáo viên phải từ 1-10 (hiện tại: {teacher_rate})")
                        error_count += 1
                        continue
                    
                    if not (1 <= class_rate <= 10):
                        errors.append(f"Dòng {row_num}: Điểm lớp học phải từ 1-10 (hiện tại: {class_rate})")
                        error_count += 1
                        continue
                    
                    # Tìm đối tượng trong database
                    try:
                        student = hoc_vien.objects.get(student_id=student_id)
                        print(f"Found student: {student.full_name}")
                    except hoc_vien.DoesNotExist:
                        errors.append(f"Dòng {row_num}: Không tìm thấy học viên với ID {student_id}")
                        error_count += 1
                        continue
                    
                    try:
                        teacher_obj = teacher.objects.get(teacher_id=teacher_id)
                        print(f"Found teacher: {teacher_obj.full_name}")
                    except teacher.DoesNotExist:
                        errors.append(f"Dòng {row_num}: Không tìm thấy giáo viên với ID {teacher_id}")
                        error_count += 1
                        continue
                    
                    try:
                        class_obj = clazz.objects.get(class_id=class_id)
                        print(f"Found class: {class_obj.class_name}")
                    except clazz.DoesNotExist:
                        errors.append(f"Dòng {row_num}: Không tìm thấy lớp học với ID {class_id}")
                        error_count += 1
                        continue
                    
                    # Kiểm tra feedback đã tồn tại chưa
                    existing_feedback = feedback.objects.filter(
                        student=student,
                        teacher=teacher_obj,
                        class_obj=class_obj
                    ).first()
                    
                    if existing_feedback:
                        # Cập nhật feedback cũ
                        existing_feedback.teacher_rate = Decimal(str(teacher_rate))
                        existing_feedback.class_rate = Decimal(str(class_rate))
                        existing_feedback.save()
                        print(f"Updated existing feedback {existing_feedback.id_feedback}")
                    else:
                        # Tạo feedback mới
                        new_feedback = feedback.objects.create(
                            student=student,
                            teacher=teacher_obj,
                            class_obj=class_obj,
                            teacher_rate=Decimal(str(teacher_rate)),
                            class_rate=Decimal(str(class_rate))
                        )
                        print(f"Created new feedback {new_feedback.id_feedback}")
                    
                    success_count += 1
                    
                except Exception as e:
                    error_msg = f"Dòng {row_num}: {str(e)}"
                    errors.append(error_msg)
                    error_count += 1
                    print(f"Error in row {row_num}: {error_msg}")
                    print(f"Traceback: {traceback.format_exc()}")
            
            # Hiển thị kết quả
            if success_count > 0:
                messages.success(request, f'Import thành công {success_count} đánh giá!')
            
            if error_count > 0:
                error_msg = f'Có {error_count} lỗi xảy ra:\n' + '\n'.join(errors[:5])
                if len(errors) > 5:
                    error_msg += f'\n... và {len(errors) - 5} lỗi khác'
                messages.error(request, error_msg)
            
            if success_count == 0 and error_count == 0:
                messages.warning(request, 'Không có dữ liệu nào được xử lý. Vui lòng kiểm tra lại file.')
                
        except Exception as e:
            error_msg = f'Lỗi đọc file: {str(e)}'
            print(f"File reading error: {error_msg}")
            print(f"Traceback: {traceback.format_exc()}")
            messages.error(request, error_msg)
    else:
        messages.error(request, 'Vui lòng chọn file để import.')
    
    return redirect('core:feedback_list')

@login_required
def feedback_view(request, pk):
    feedback_obj = get_object_or_404(feedback, pk=pk)
    return render(request, 'core/feedback_detail.html', {'feedback': feedback_obj})

@login_required
def feedback_delete(request, pk):
    feedback_obj = get_object_or_404(feedback, pk=pk)
    if request.method == 'POST':
        feedback_obj.delete()
        messages.success(request, 'Đánh giá đã được xóa thành công')
    return redirect('core:feedback_list')

#-------------------------------
# Nhan Vien Management Views
#-------------------------------
@login_required
def nhanvien_list(request):
    search_query = request.GET.get('search', '')
    nhanviens = nhan_vien.objects.all()
    
    if search_query:
        nhanviens = nhanviens.filter(
            Q(full_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(sdt__icontains=search_query) |
            Q(nv_id__icontains=search_query)
        )
    
    # Phân trang - 10 nhân viên mỗi trang
    paginator = Paginator(nhanviens, 10)
    page = request.GET.get('page')
    
    try:
        nhanviens_page = paginator.page(page)
    except PageNotAnInteger:
        nhanviens_page = paginator.page(1)
    except EmptyPage:
        nhanviens_page = paginator.page(paginator.num_pages)
    
    context = {
        'nhanviens': nhanviens_page,
        'search_query': search_query,
    }
    return render(request, 'nhanvien/nhanvien_list.html', context)

@login_required
def nhanvien_create(request):
    if request.method == 'POST':
        form = NhanVienForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Thêm nhân viên thành công!')
            return redirect('core:nhanvien_list')
    else:
        form = NhanVienForm()
    return render(request, 'nhanvien/nhanvien_form.html', {'form': form, 'action': 'Thêm'})

@login_required
def nhanvien_edit(request, pk):
    nv = get_object_or_404(nhan_vien, pk=pk)
    if request.method == 'POST':
        form = NhanVienForm(request.POST, instance=nv)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cập nhật nhân viên thành công!')
            return redirect('core:nhanvien_list')
    else:
        form = NhanVienForm(instance=nv)
    return render(request, 'nhanvien/nhanvien_form.html', {'form': form, 'action': 'Sửa'})

@login_required
def nhanvien_delete(request, pk):
    nv = get_object_or_404(nhan_vien, pk=pk)
    if request.method == 'POST':
        nv.delete()
        messages.success(request, 'Xóa nhân viên thành công!')
        return redirect('core:nhanvien_list')
    return render(request, 'nhanvien/nhanvien_confirm_delete.html', {'nhanvien': nv})


#-------------------------------
# Class Type Management Views
#-------------------------------
@login_required
def class_type_list(request):
    types = class_type.objects.all()
    
    # Phân trang - 10 loại lớp mỗi trang
    paginator = Paginator(types, 10)
    page = request.GET.get('page')
    
    try:
        types_page = paginator.page(page)
    except PageNotAnInteger:
        types_page = paginator.page(1)
    except EmptyPage:
        types_page = paginator.page(paginator.num_pages)
    
    return render(request, 'class_type/class_type_list.html', {'types': types_page})

@login_required
def class_type_create(request):
    if request.method == 'POST':
        form = ClassTypeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Thêm loại lớp thành công!')
            return redirect('core:class_type_list')
    else:
        form = ClassTypeForm()
    return render(request, 'class_type/class_type_form.html', {'form': form, 'action': 'Thêm'})

@login_required
def class_type_edit(request, pk):
    ct = get_object_or_404(class_type, pk=pk)
    if request.method == 'POST':
        form = ClassTypeForm(request.POST, instance=ct)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cập nhật loại lớp thành công!')
            return redirect('core:class_type_list')
    else:
        form = ClassTypeForm(instance=ct)
    return render(request, 'class_type/class_type_form.html', {'form': form, 'action': 'Sửa'})

@login_required
def class_type_delete(request, pk):
    ct = get_object_or_404(class_type, pk=pk)
    if request.method == 'POST':
        ct.delete()
        messages.success(request, 'Xóa loại lớp thành công!')
        return redirect('core:class_type_list')
    return render(request, 'class_type/class_type_confirm_delete.html', {'class_type': ct})
